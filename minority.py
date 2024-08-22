from hashlib import new
import pandas as pd
from numpy.random import rand
from numpy.random import choice
from numpy import asarray
from numpy import clip
from numpy import argmin
from numpy import min
from numpy import around
from codecs import ignore_errors


# import module
import openpyxl
import random
import numpy as np
  


from dask.dataframe.core import DataFrame
#import excel file
file=pd.ExcelFile('Storm_reduced_features_withdata.xlsx')
df1=file.parse('Sheet1')
print('Number of rows and columns in excel',df1.shape) 
#create a copy of data
#Independent Variables
min =df1 [ df1.Output == 1]
print('Number of rows and columns in minority instances',min.shape) 
maj=df1 [df1.Output == 0]
print('Number of rows and columns in majority instances',maj.shape) 
#df1=df1[min]
df2=min.drop(columns=['Output'])
#print(df2)
df2.to_excel (r'minority_instances.xlsx', index = False, header=True)
#maj.to_excel(r'majority_instances.xlsx', index=False, header=True)
#Dependent Variable
#y=df1[['Output']]
file=pd.ExcelFile('minority_instances.xlsx')
df1=file.parse('Sheet1')
print('Number of rows and columns in minority file',df1.shape)
df_max_column= df1.copy()
#apply normalisation (min-max)
for column in df_max_column.columns:
    df_max_column[column]= (df_max_column[column]- df_max_column[column].min())/(df_max_column[column].max()-df_max_column[column].min())
    df_max_column[column]=round(df_max_column[column],2)
# print(df_max_column)
# Max_values=df_max_column.max(numeric_only= True)
# print(Max_values)
# Min_values=df_max_column.min(numeric_only=True)
# print(Min_values)
df_max_column.to_excel (r'min_max_normalised.xlsx', index = False, header=True) 

# load excel with its path
wrkbk = openpyxl.load_workbook("min_max_normalised.xlsx")
sh = wrkbk.active

F=0.3
cr=0.9
iter=10
obj_all= []

def crossover(mutant_vector,column):
    sum=0.0
    # generate a uniform random value for every dimension
    p = random.uniform(0,1)
    #print('value of p',p)
    # generate trial vector by binomial crossover
    if p < cr:
        trial=mutant_vector
    else:
        for j in range(2,sh.max_row+1):
            sum=sum+sh.cell(j,column).value
        avera=float(sum/ (sh.max_row-1))
        trial=avera            
    return trial  
matrix=[]
for k in range(iter):
    row=[]
    for i in range(1, sh.max_column+1):
        column = i
        row1=random.randint(2,sh.max_row+1)
        value1=sh.cell(row1,column).value
        value1=float(0 if value1 is None else value1)
        row2=random.randint(2,sh.max_row+1)
        value2=sh.cell(row2,column).value
        value2=float(0 if value2 is None else value2)
        row3=random.randint(2,sh.max_row+1)
        value3=sh.cell(row2,column).value
        value3=float(0 if value3 is None else value3)
        mutant_vector=float(value1 + F * (value2 - value3))
        #print(column , mutant_vector)
        trial = float(crossover(mutant_vector,column))
        #print (column,trial)
        row.append(trial)
    matrix.append(row)
#print(matrix)
#print(len(matrix))
#print(len(matrix[0]))
dfexc=pd.DataFrame(matrix).T
#dfexc.to_excel (r'matrix.xlsx', index = False, header=True) 
av_row=dfexc.mean(axis=1)
av_row=round(av_row,2)
#print(av_row)  
#result=df_max_column.mul(av_row,axis=0)
#result.to_excel (r'C:\Users\Harguneet Kaur\eclipse-workspace\Optimization\result.xlsx', index = False, header=True) 
# df_av=av_row.describe()
# print(df_av)
#multiplying by weights
i=0
df_weighted = df_max_column.copy()
for column in df_weighted.columns:
    df_weighted[column]*=av_row[i]
    i+=1

#adding complexity column
complexity = df_weighted.sum(axis=1)
df_weighted['Complexity'] = complexity

#sorting dataframe according to complexity
df_weighted_sorted = df_weighted.sort_values(['Complexity'])
#df_weighted_sorted.to_excel (r'weighted_sorted.xlsx', index = False, header=True)
#making new data frame
mid = 0 #variable to check if while loop breaks before adding new instances back to sorted df
new_min = df_weighted_sorted.shape[0] #variable for number of minority instances
while new_min<=maj.shape[0]:
    new_df = pd.DataFrame(columns = df_weighted.columns)
    #creating new instances
    for i in range(len(df_weighted_sorted)-1):
        new_df.loc[len(new_df)] = df_weighted_sorted[i:i+2].mean(axis=0)
        new_min += 1
        if new_min == maj.shape[0]:
            mid = 1
            break
    if mid == 1:
        break
    #adding new instances back to sorted df if loop doesn't break midway
    for i in range(len(new_df)):
        df_weighted_sorted.loc[len(df_weighted_sorted)] = new_df.loc[i]
    df_weighted_sorted = df_weighted_sorted.sort_values(['Complexity'])
# adding new instances back to sorted df if loop breaks midway
if mid == 1:
    for i in range(len(new_df)):
        df_weighted_sorted.loc[len(df_weighted_sorted)] = new_df.loc[i]
    df_weighted_sorted = df_weighted_sorted.sort_values(['Complexity'])
df_weighted_sorted.to_excel (r'final_weighted.xlsx', index = False, header=True)  

file=pd.ExcelFile('final_weighted.xlsx')
dfM=file.parse('Sheet1')
dfF=dfM.drop(columns=['Complexity'])
Output_value=1
dfF['Output']=Output_value
dfF.to_excel (r'Final_Minority_Output.xlsx', index = False, header=True)
result = maj.append(dfF)
result.to_csv (r'Balanced_File.csv', index = False, header=True)



