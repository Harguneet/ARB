import pandas as pd
from numpy.random import rand
from numpy.random import choice
from numpy import asarray
from numpy import clip
from numpy import argmin
from numpy import min
from numpy import around
from codecs import ignore_errors


# import module
import openpyxl
import random
import numpy as np
  


from dask.dataframe.core import DataFrame
#import excel file
file=pd.ExcelFile('cassandra_reducedfeatures_withdata.xlsx')
df1=file.parse('Sheet1')
print('Number of rows and columns in excel',df1.shape) 
#create a copy of data
#Independent Variables
min =df1 [ df1.Output == 1]
print('Number of rows and columns in minority instances',min.shape) 
maj=df1 [df1.Output == 0]
print('Number of rows and columns in majority instances',maj.shape) 
#df1=df1[min]
df2=min.drop(columns=['Output'])
#print(df2)
df2.to_excel (r'C:\Users\Harguneet Kaur\eclipse-workspace\Optimization\minority_instances.xlsx', index = False, header=True)
#Dependent Variable
#y=df1[['Output']]
file=pd.ExcelFile('minority_instances.xlsx')
df1=file.parse('Sheet1')
print('Number of rows and columns in minority file',df1.shape)
df_max_column= df1.copy()

#apply normalisation (min-max)
for column in df_max_column.columns:
    df_max_column[column]= (df_max_column[column]- df_max_column[column].min())/(df_max_column[column].max()-df_max_column[column].min())
    df_max_column[column]=round(df_max_column[column],2)
# print(df_max_column)
# Max_values=df_max_column.max(numeric_only= True)
# print(Max_values)
# Min_values=df_max_column.min(numeric_only=True)
# print(Min_values)
df_max_column.to_excel (r'C:\Users\Harguneet Kaur\eclipse-workspace\Optimization\min_max_normalised.xlsx', index = False, header=True) 

# load excel with its path
wrkbk = openpyxl.load_workbook("min_max_normalised.xlsx")
sh = wrkbk.active

F=0.3
cr=0.9
iter=10
obj_all= []

def crossover(mutant_vector,column):
    sum=0.0
    # generate a uniform random value for every dimension
    p = random.uniform(0,1)
    #print('value of p',p)
    # generate trial vector by binomial crossover
    if p < cr:
        trial=mutant_vector
    else:
        for j in range(2,sh.max_row+1):
            sum=sum+sh.cell(j,column).value
        avera=float(sum/ (sh.max_row-1))
        trial=avera            
    return trial  
matrix=[]
for k in range(iter):
    row=[]
    for i in range(1, sh.max_column+1):
        column = i
        row1=random.randint(2,sh.max_row+1)
        value1=sh.cell(row1,column).value
        row2=random.randint(2,sh.max_row+1)
        value2=sh.cell(row2,column).value
        row3=random.randint(2,sh.max_row+1)
        value3=sh.cell(row2,column).value
        mutant_vector=float(value1 + F * (value2 - value3))
        #print(column , mutant_vector)
        trial = float(crossover(mutant_vector,column))
        #print (column,trial)
        row.append(trial)
    matrix.append(row)
print(matrix)
print(len(matrix))
print(len(matrix[0]))
dfexc=pd.DataFrame(matrix).T
#dfexc.to_excel (r'C:\Users\Harguneet Kaur\eclipse-workspace\Optimization\matrix.xlsx', index = False, header=True) 
av_row=dfexc.mean(axis=1)
av_row=round(av_row,2)
print(av_row)  
#result=df_max_column.mul(av_row,axis=0)
#result.to_excel (r'C:\Users\Harguneet Kaur\eclipse-workspace\Optimization\result.xlsx', index = False, header=True) 
# df_av=av_row.describe()
# print(df_av)
#multiplying by weights
i=0
df_weighted = df_max_column.copy()
for column in df_weighted.columns:
    df_weighted[column]*=av_row[i]
    i+=1

#adding complexity column
complexity = df_weighted.sum(axis=1)
df_weighted['Complexity'] = complexity

#sorting dataframe according to complexity
df_weighted_sorted = df_weighted.sort_values(['Complexity'])

#making new data frame
df = df_weighted_sorted.T
new_df = pd.DataFrame(columns=df_weighted.columns)

#creating new instances
for i in range(len(df_weighted)-1):
    new_df.loc[len(new_df)]=(df[i]+df[i+1])/2

#adding new instances back to sorted df
for i in range(len(new_df)):
    df_weighted_sorted.loc[(2*i)+0.5] = new_df.loc[i]
df_weighted_sorted.to_excel (r'final_weighted.xlsx', index = False, header=True)  

